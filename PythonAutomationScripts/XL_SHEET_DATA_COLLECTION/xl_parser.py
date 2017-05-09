import sys, os
import json

from point import Point
from path import Path
from db_helper import Db_helper

#find files in current directory
dir_path = os.path.dirname(os.path.realpath(__file__))
sys.path.append(dir_path+"\et_xmlfile-1.0.1")
sys.path.append(dir_path+"\jdcal-1.3")
sys.path.append(dir_path+"\openpyxl-2.4.4")
import openpyxl
class Xl_parser(object):
    def __init__(self):
        pass

    def parse_wb(self, wb):
        sheet_names = wb.get_sheet_names()
        paths = []
        points = []
        for sheet_name in sheet_names:
            if(sheet_name == "Paths"):
                #insert into path table for building name
                paths += self.parse_paths(wb.get_sheet_by_name(sheet_name))
            else:
                #insert into point table for building name
                points += self.parse_points(wb.get_sheet_by_name(sheet_name))

        return (points, paths)
    def parse_points(self, sheet):
        row = 2
        points = []
        while(sheet.cell(row = row, column = 1).value != None):
            #print sheet.cell(row = row, column = 1).value
            #parse row into point
            name = sheet.cell(row = row, column = 1).value
            entry = sheet.cell(row = row, column = 2).value
            lat = sheet.cell(row = row, column = 3).value
            lng = sheet.cell(row = row, column = 4).value
            col = 5
            next_pts = []
            while(sheet.cell(row = row, column = col).value != None):
                next_pts.append(sheet.cell(row = row, column = col).value)
                print "  ",sheet.cell(row = row, column = col).value
                col += 1
            next_pts = json.dumps({'next_pts':next_pts})
            points.append(Point(name, lat, lng, next_pts, entry))
            row += 1
        return points

    def parse_paths(self, sheet):
        paths = []
        row = 2
        while(sheet.cell(row = row, column = 1).value != None):
            #print sheet.cell(row = row, column = 1).value
            #parse row into path
            start = sheet.cell(row = row, column = 1).value
            end = sheet.cell(row = row, column = 2).value
            name = sheet.cell(row = row, column = 3).value
            paths.append(Path(start, end, name))
            row += 1

        return paths

def main():
    #print 'Number of args:',len(sys.argv),'\n'
    #print 'arg list: ',str(sys.argv),'\n'
    if len(sys.argv) != 2:
        print 'Usage: python program.py building_name.xlsx'
        return
    wb = openpyxl.load_workbook(sys.argv[1])
    db_name = sys.argv[1].replace(".xlsx","").upper()
    print db_name
    xl = Xl_parser()
    db = Db_helper(db_name)
    (points, paths) = xl.parse_wb(wb)
    #db.purge_paths()
    #db.purge_points()
    #db.insert_points(points)
    #db.insert_paths(paths)
    #db.print_table('PATHS')
    db.print_table('POINTS')



if __name__ == '__main__':
    main()
